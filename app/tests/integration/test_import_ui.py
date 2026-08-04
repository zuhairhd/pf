"""Import Center UI tests (IMP-703).

Covers the browser-facing Import Center, upload forms (CSV/Excel/SMS),
preview/confirm/cancel flows, account-picker visibility, read-only safety,
and tenant/RLS isolation. Every route under test is a thin wrapper around the
existing ImportService methods already covered by test_imports.py -- these
tests exercise the UI layer, not the parsers themselves.

Only synthetic/fake data is used -- no real bank statements or SMS messages.
"""

from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest
from sqlalchemy import select

from app.core.rls import set_tenant_context_async
from app.models import (
    Account,
    Budget,
    Family,
    FamilyRole,
    Goal,
    ImportJob,
    JournalEntry,
)
from app.tests.helpers import (
    assert_rls_enabled,
    auth_headers_for,
    count_rows,
    create_test_account,
    create_test_family_member,
    create_test_organization,
    create_test_user,
)


def _csv_bytes(text: str) -> bytes:
    return text.encode("utf-8")


def _file_tuple(name: str, content: bytes, content_type: str = "text/csv"):
    return (name, io.BytesIO(content), content_type)


def _build_xlsx(rows: list[list], headers: list[str], sheet_name: str = "Sheet1") -> bytes:
    """Build a fake .xlsx workbook in memory -- no real bank data is used."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


VALID_CSV = (
    "Date,Description,Amount\n"
    "2026-07-01,Salary deposit,1500.000\n"
    "2026-07-02,Grocery store,-45.500\n"
)

MIXED_CSV = (
    "Date,Description,Amount\n"
    "2026-07-01,Coffee,-2.500\n"
    "2026-07-01,Coffee,-2.500\n"  # duplicate of the row above
    ",Missing date,-5.000\n"  # invalid
)


async def _setup_accounts(client, auth_headers) -> dict:
    accounts = [
        {"code": "BANK", "name": "Bank Muscat", "account_type": "Asset", "is_bank_account": True},
        {"code": "SAL", "name": "Salary", "account_type": "Income"},
        {"code": "GRO", "name": "Food & Groceries", "account_type": "Expense"},
    ]
    created = {}
    for account in accounts:
        response = await client.post("/accounts/", json=account, headers=auth_headers)
        assert response.status_code == 200, response.text
        created[account["code"]] = response.json()
    return created


# ---------------------------------------------------------------------------
# Import Center
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_import_center_requires_auth(client):
    response = await client.get("/imports")
    assert response.status_code in (401, 403)


@pytest.mark.integration
@pytest.mark.anyio
async def test_import_center_shows_method_cards(client, auth_headers):
    response = await client.get("/imports", headers=auth_headers)
    assert response.status_code == 200, response.text
    assert "CSV File" in response.text
    assert "Excel File" in response.text
    assert "Paste SMS" in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_import_center_empty_state_renders(client, db, unique):
    org = await create_test_organization(db, name=unique("EmptyOrg"), slug=unique("empty-org"))
    user, password = await create_test_user(db, org)
    headers = await auth_headers_for(client, user.email, password)

    response = await client.get("/imports", headers=headers)
    assert response.status_code == 200, response.text
    assert "No import jobs yet" in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_import_history_shows_jobs(client, auth_headers):
    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text

    response = await client.get("/imports", headers=auth_headers)
    assert response.status_code == 200, response.text
    assert "statement.csv" in response.text


# ---------------------------------------------------------------------------
# CSV UI
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_csv_form_renders(client, auth_headers):
    response = await client.get("/imports/partials/csv-form", headers=auth_headers)
    assert response.status_code == 200, response.text
    assert 'name="file"' in response.text
    assert 'name="mapping"' in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_csv_upload_via_ui_creates_preview_job(client, auth_headers):
    response = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    assert "Preview" in response.text
    assert "statement.csv" in response.text
    assert "import-preview-panel" in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_csv_invalid_upload_shows_safe_error(client, auth_headers):
    response = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("empty.csv", b"")},
        headers=auth_headers,
    )
    assert response.status_code == 400
    assert "empty" in response.text.lower()
    # Safe inline error, not a raw 500/traceback.
    assert "Traceback" not in response.text


# ---------------------------------------------------------------------------
# Excel UI
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_excel_form_renders(client, auth_headers):
    response = await client.get("/imports/partials/excel-form", headers=auth_headers)
    assert response.status_code == 200, response.text
    assert 'name="sheet_name"' in response.text
    assert 'name="default_currency"' in response.text
    assert 'name="default_account_id"' in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_excel_upload_via_ui_creates_preview_job(client, auth_headers):
    content = _build_xlsx(
        [[date(2026, 7, 1), "Salary deposit", 1500.0]],
        ["Date", "Description", "Amount"],
    )
    response = await client.post(
        "/imports/ui/excel",
        files={
            "file": (
                "statement.xlsx",
                io.BytesIO(content),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    assert "Preview" in response.text
    assert "import-preview-panel" in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_excel_upload_sheet_name_works(client, auth_headers):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["irrelevant"])
    ws2 = wb.create_sheet("Statement")
    ws2.append(["Date", "Description", "Amount"])
    ws2.append([date(2026, 7, 3), "Bonus payment", 250.0])
    buf = io.BytesIO()
    wb.save(buf)

    response = await client.post(
        "/imports/ui/excel",
        data={"sheet_name": "Statement"},
        files={
            "file": (
                "multi_sheet.xlsx",
                io.BytesIO(buf.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    assert "Bonus payment" in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_excel_upload_rejects_unsupported_xls_extension(client, auth_headers):
    response = await client.post(
        "/imports/ui/excel",
        files={"file": _file_tuple("statement.xls", b"not a real xls file", "application/vnd.ms-excel")},
        headers=auth_headers,
    )
    assert response.status_code == 400
    assert "xlsx" in response.text.lower()
    assert "Traceback" not in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_excel_upload_corrupted_workbook_shows_safe_error(client, auth_headers):
    response = await client.post(
        "/imports/ui/excel",
        files={
            "file": (
                "statement.xlsx",
                io.BytesIO(b"this is not a real workbook"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert response.status_code == 400
    assert "Traceback" not in response.text


# ---------------------------------------------------------------------------
# SMS UI
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_sms_form_renders(client, auth_headers):
    response = await client.get("/imports/partials/sms-form", headers=auth_headers)
    assert response.status_code == 200, response.text
    assert 'name="sms_text"' in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_sms_parse_via_ui_creates_preview_job(client, auth_headers):
    text = (
        "Bank Muscat: Your account ****1234 has been debited OMR 45.000 "
        "on 01-JUL-2026 at CARREFOUR. Avl Bal OMR 1234.567"
    )
    response = await client.post(
        "/imports/ui/sms",
        data={"sms_text": text},
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    assert "Preview" in response.text
    assert "SMS" in response.text.upper()


@pytest.mark.integration
@pytest.mark.anyio
async def test_sms_invalid_shows_invalid_row_preview(client, auth_headers):
    response = await client.post(
        "/imports/ui/sms",
        data={"sms_text": "This is just a random text message with no bank info."},
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    assert "Invalid" in response.text


# ---------------------------------------------------------------------------
# Preview / confirm / cancel
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_preview_page_requires_auth(client, auth_headers, test_user, db, tenant_context):
    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text

    await tenant_context(test_user.organization_id)
    result = await db.execute(
        select(ImportJob).where(ImportJob.tenant_id == test_user.organization_id)
    )
    job = result.scalars().first()
    assert job is not None

    page_response = await client.get(f"/imports/ui/{job.id}/preview")
    assert page_response.status_code in (401, 403)

    partial_response = await client.get(f"/imports/partials/{job.id}/preview")
    assert partial_response.status_code in (401, 403)


@pytest.mark.integration
@pytest.mark.anyio
async def test_preview_shows_valid_invalid_duplicate_rows(client, auth_headers):
    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("mixed.csv", _csv_bytes(MIXED_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text
    assert "Valid" in upload.text
    assert "Invalid" in upload.text
    assert "Duplicate" in upload.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_confirm_button_posts_valid_rows(client, auth_headers, test_user, db, tenant_context):
    accounts = await _setup_accounts(client, auth_headers)

    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text

    await tenant_context(test_user.organization_id)
    result = await db.execute(
        select(ImportJob).where(ImportJob.tenant_id == test_user.organization_id)
    )
    job = result.scalars().first()
    assert job is not None

    confirm = await client.post(
        f"/imports/ui/{job.id}/confirm",
        data={
            "bank_account_id": accounts["BANK"]["id"],
            "default_income_account_id": accounts["SAL"]["id"],
            "default_expense_account_id": accounts["GRO"]["id"],
        },
        headers=auth_headers,
    )
    assert confirm.status_code == 200, confirm.text
    assert "Completed" in confirm.text
    assert "Imported 2 row" in confirm.text

    result = await db.execute(
        select(JournalEntry)
        .where(JournalEntry.tenant_id == test_user.organization_id)
        .where(JournalEntry.source == "import")
    )
    entries = result.scalars().all()
    assert len(entries) == 2


@pytest.mark.integration
@pytest.mark.anyio
async def test_cancel_button_cancels_job(client, auth_headers, test_user, db, tenant_context):
    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text

    await tenant_context(test_user.organization_id)
    result = await db.execute(
        select(ImportJob).where(ImportJob.tenant_id == test_user.organization_id)
    )
    job = result.scalars().first()
    assert job is not None

    cancel = await client.post(f"/imports/ui/{job.id}/cancel", headers=auth_headers)
    assert cancel.status_code == 200, cancel.text
    assert "Cancelled" in cancel.text

    status_response = await client.get(f"/imports/{job.id}", headers=auth_headers)
    assert status_response.json()["status"] == "cancelled"

    result = await db.execute(
        select(JournalEntry).where(JournalEntry.tenant_id == test_user.organization_id)
    )
    assert len(result.scalars().all()) == 0


@pytest.mark.integration
@pytest.mark.anyio
async def test_duplicate_rows_are_not_posted(client, auth_headers, test_user, db, tenant_context):
    accounts = await _setup_accounts(client, auth_headers)

    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("mixed.csv", _csv_bytes(MIXED_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text

    await tenant_context(test_user.organization_id)
    result = await db.execute(
        select(ImportJob).where(ImportJob.tenant_id == test_user.organization_id)
    )
    job = result.scalars().first()
    assert job.valid_rows == 1
    assert job.duplicate_rows == 1
    assert job.invalid_rows == 1

    confirm = await client.post(
        f"/imports/ui/{job.id}/confirm",
        data={
            "bank_account_id": accounts["BANK"]["id"],
            "default_expense_account_id": accounts["GRO"]["id"],
        },
        headers=auth_headers,
    )
    assert confirm.status_code == 200, confirm.text

    result = await db.execute(
        select(JournalEntry)
        .where(JournalEntry.tenant_id == test_user.organization_id)
        .where(JournalEntry.source == "import")
    )
    assert len(result.scalars().all()) == 1


# ---------------------------------------------------------------------------
# Account picker visibility
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_account_picker_hides_inaccessible_private_accounts(client, db, unique):
    org = await create_test_organization(db, name=unique("VisOrg"), slug=unique("vis-org"))
    await set_tenant_context_async(db, org.id)
    head, _ = await create_test_user(db, org, email=unique("head") + "@example.com", role="owner")
    viewer, viewer_password = await create_test_user(
        db, org, email=unique("viewer") + "@example.com", role="viewer"
    )

    family = Family(tenant_id=org.id, name=unique("Family"), currency="OMR")
    db.add(family)
    await db.flush()
    await db.refresh(family)
    await create_test_family_member(db, family.id, org.id, head, FamilyRole.HEAD.value)
    await create_test_family_member(db, family.id, org.id, viewer, FamilyRole.VIEWER.value)
    await db.commit()

    private_account = await create_test_account(
        db,
        org.id,
        code=unique("PRIV"),
        name="Heads Private Stash",
        account_type="Asset",
        visibility="private",
        owner_user_id=head.id,
    )

    viewer_headers = await auth_headers_for(client, viewer.email, viewer_password)
    response = await client.get("/imports/partials/excel-form", headers=viewer_headers)
    assert response.status_code == 200, response.text
    assert private_account.code not in response.text
    assert "Heads Private Stash" not in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_account_picker_never_shows_cross_tenant_accounts(client, db, unique):
    org_a = await create_test_organization(db, name=unique("Org A"), slug=unique("org-a"))
    org_b = await create_test_organization(db, name=unique("Org B"), slug=unique("org-b"))
    user_a, password_a = await create_test_user(db, org_a)

    await set_tenant_context_async(db, org_b.id)
    b_account = await create_test_account(
        db, org_b.id, code=unique("BACC"), name="Org B Secret Account", account_type="Asset"
    )
    await db.commit()

    headers_a = await auth_headers_for(client, user_a.email, password_a)
    response = await client.get("/imports/partials/excel-form", headers=headers_a)
    assert response.status_code == 200, response.text
    assert b_account.code not in response.text
    assert "Org B Secret Account" not in response.text


@pytest.mark.integration
@pytest.mark.anyio
async def test_inaccessible_account_submission_rejected(client, db, unique):
    org = await create_test_organization(db, name=unique("PermOrg"), slug=unique("perm-org"))
    await set_tenant_context_async(db, org.id)
    head, _ = await create_test_user(db, org, email=unique("head") + "@example.com", role="owner")
    viewer, viewer_password = await create_test_user(
        db, org, email=unique("viewer") + "@example.com", role="viewer"
    )

    family = Family(tenant_id=org.id, name=unique("Family"), currency="OMR")
    db.add(family)
    await db.flush()
    await db.refresh(family)
    await create_test_family_member(db, family.id, org.id, head, FamilyRole.HEAD.value)
    await create_test_family_member(db, family.id, org.id, viewer, FamilyRole.VIEWER.value)
    await db.commit()

    private_bank = await create_test_account(
        db,
        org.id,
        code=unique("PBANK"),
        name="Heads Private Bank",
        account_type="Asset",
        visibility="private",
        owner_user_id=head.id,
    )
    expense = await create_test_account(
        db, org.id, code=unique("EXP"), name="Groceries", account_type="Expense"
    )

    viewer_headers = await auth_headers_for(client, viewer.email, viewer_password)
    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=viewer_headers,
    )
    assert upload.status_code == 200, upload.text

    await set_tenant_context_async(db, org.id)
    result = await db.execute(select(ImportJob).where(ImportJob.tenant_id == org.id))
    job = result.scalars().first()
    assert job is not None

    confirm = await client.post(
        f"/imports/ui/{job.id}/confirm",
        data={
            "bank_account_id": private_bank.id,
            "default_expense_account_id": expense.id,
        },
        headers=viewer_headers,
    )
    assert confirm.status_code == 200, confirm.text
    assert "Imported 0 row" in confirm.text


# ---------------------------------------------------------------------------
# Read-only safety
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_opening_import_center_creates_no_financial_records(
    client, auth_headers, test_user, db, tenant_context
):
    await tenant_context(test_user.organization_id)
    before = {
        "journal_entries": await count_rows(
            db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
        ),
        "accounts": await count_rows(db, Account, Account.tenant_id == test_user.organization_id),
        "budgets": await count_rows(db, Budget, Budget.tenant_id == test_user.organization_id),
        "goals": await count_rows(db, Goal, Goal.tenant_id == test_user.organization_id),
    }

    response = await client.get("/imports", headers=auth_headers)
    assert response.status_code == 200

    await tenant_context(test_user.organization_id)
    after = {
        "journal_entries": await count_rows(
            db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
        ),
        "accounts": await count_rows(db, Account, Account.tenant_id == test_user.organization_id),
        "budgets": await count_rows(db, Budget, Budget.tenant_id == test_user.organization_id),
        "goals": await count_rows(db, Goal, Goal.tenant_id == test_user.organization_id),
    }
    assert before == after


@pytest.mark.integration
@pytest.mark.anyio
async def test_previewing_import_creates_no_journal_entries(
    client, auth_headers, test_user, db, tenant_context
):
    await tenant_context(test_user.organization_id)
    before = await count_rows(
        db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
    )

    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200, upload.text

    await tenant_context(test_user.organization_id)
    result = await db.execute(
        select(ImportJob).where(ImportJob.tenant_id == test_user.organization_id)
    )
    job = result.scalars().first()

    preview = await client.get(f"/imports/ui/{job.id}/preview", headers=auth_headers)
    assert preview.status_code == 200
    partial = await client.get(f"/imports/partials/{job.id}/preview", headers=auth_headers)
    assert partial.status_code == 200

    await tenant_context(test_user.organization_id)
    after = await count_rows(
        db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
    )
    assert before == after


@pytest.mark.integration
@pytest.mark.anyio
async def test_only_confirm_creates_journal_entries(
    client, auth_headers, test_user, db, tenant_context
):
    accounts = await _setup_accounts(client, auth_headers)

    await tenant_context(test_user.organization_id)
    before = await count_rows(
        db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
    )

    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=auth_headers,
    )
    assert upload.status_code == 200

    await tenant_context(test_user.organization_id)
    result = await db.execute(
        select(ImportJob).where(ImportJob.tenant_id == test_user.organization_id)
    )
    job = result.scalars().first()

    await tenant_context(test_user.organization_id)
    after_upload = await count_rows(
        db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
    )
    assert after_upload == before

    confirm = await client.post(
        f"/imports/ui/{job.id}/confirm",
        data={
            "bank_account_id": accounts["BANK"]["id"],
            "default_income_account_id": accounts["SAL"]["id"],
            "default_expense_account_id": accounts["GRO"]["id"],
        },
        headers=auth_headers,
    )
    assert confirm.status_code == 200, confirm.text

    await tenant_context(test_user.organization_id)
    after_confirm = await count_rows(
        db, JournalEntry, JournalEntry.tenant_id == test_user.organization_id
    )
    assert after_confirm == before + 2


# ---------------------------------------------------------------------------
# Tenant / RLS isolation
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.anyio
async def test_tenant_a_cannot_see_tenant_b_import_jobs(client, db, unique):
    org_a = await create_test_organization(db, name=unique("Org A"), slug=unique("org-a"))
    org_b = await create_test_organization(db, name=unique("Org B"), slug=unique("org-b"))
    user_a, password_a = await create_test_user(db, org_a)
    user_b, password_b = await create_test_user(db, org_b)

    headers_a = await auth_headers_for(client, user_a.email, password_a)
    headers_b = await auth_headers_for(client, user_b.email, password_b)

    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=headers_a,
    )
    assert upload.status_code == 200, upload.text

    await set_tenant_context_async(db, org_a.id)
    result = await db.execute(select(ImportJob).where(ImportJob.tenant_id == org_a.id))
    job = result.scalars().first()
    assert job is not None

    response_b = await client.get(f"/imports/ui/{job.id}/preview", headers=headers_b)
    assert response_b.status_code == 404

    partial_b = await client.get(f"/imports/partials/{job.id}/preview", headers=headers_b)
    assert partial_b.status_code == 404


@pytest.mark.integration
@pytest.mark.anyio
async def test_tenant_a_cannot_confirm_or_cancel_tenant_b_jobs(client, db, unique):
    org_a = await create_test_organization(db, name=unique("Org A"), slug=unique("org-a"))
    org_b = await create_test_organization(db, name=unique("Org B"), slug=unique("org-b"))
    user_a, password_a = await create_test_user(db, org_a)
    user_b, password_b = await create_test_user(db, org_b)

    headers_a = await auth_headers_for(client, user_a.email, password_a)
    headers_b = await auth_headers_for(client, user_b.email, password_b)

    upload = await client.post(
        "/imports/ui/csv",
        files={"file": _file_tuple("statement.csv", _csv_bytes(VALID_CSV))},
        headers=headers_a,
    )
    assert upload.status_code == 200, upload.text

    await set_tenant_context_async(db, org_a.id)
    result = await db.execute(select(ImportJob).where(ImportJob.tenant_id == org_a.id))
    job = result.scalars().first()
    assert job is not None

    confirm_b = await client.post(
        f"/imports/ui/{job.id}/confirm",
        data={"bank_account_id": 1},
        headers=headers_b,
    )
    assert confirm_b.status_code == 404

    cancel_b = await client.post(f"/imports/ui/{job.id}/cancel", headers=headers_b)
    assert cancel_b.status_code == 404


@pytest.mark.integration
@pytest.mark.anyio
async def test_rls_active_on_import_and_journal_tables(db):
    await assert_rls_enabled(db, "import_jobs")
    await assert_rls_enabled(db, "imported_rows")
    await assert_rls_enabled(db, "journal_entries")
    await assert_rls_enabled(db, "journal_lines")
