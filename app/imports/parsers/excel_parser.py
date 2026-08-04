"""Excel (.xlsx) parser for the PF import module.

Reuses the same column-alias detection and duplicate-key logic as the CSV
parser (`app/imports/parsers/csv_parser.py`) so mapping and validation
behavior stay consistent across import sources. Only openpyxl's read-only,
data-only mode is used: workbook formulas and macros are never executed,
only cached cell values are read. Legacy `.xls` workbooks are not supported.
"""

from __future__ import annotations

import hashlib
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.imports.parsers.csv_parser import (
    _build_duplicate_key,
    _detect_mapping,
    _parse_date,
    _parse_decimal,
)
from app.imports.schemas import ColumnMapping


class ExcelParseError(Exception):
    """Raised when a workbook cannot be read or has no usable data."""


class ParsedExcelRow:
    """Result of parsing a single Excel row."""

    def __init__(self, row_number: int, raw_data: dict[str, Any]):
        self.row_number = row_number
        self.raw_data = raw_data
        self.parsed_data: dict[str, Any] = {}
        self.validation_errors: list[str] = []
        self.duplicate_key: Optional[str] = None
        self.status = "valid"

    def add_error(self, message: str) -> None:
        self.validation_errors.append(message)
        self.status = "invalid"


def _json_safe(value: Any) -> Any:
    """Convert an openpyxl cell value into a JSON-serializable value."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _cell_to_date(value: Any) -> Optional[date]:
    """Normalize a cell value (already-typed date or free text) to a date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _parse_date(str(value))


def _resolve_amount(raw: dict[str, Any], mapping: dict[str, str]) -> Optional[Decimal]:
    """Resolve amount from a single amount column or debit/credit columns."""
    amount_col = mapping.get("amount")
    if amount_col:
        amount = _parse_decimal(raw.get(amount_col))
        if amount is not None and amount != 0:
            return amount

    debit_col = mapping.get("debit")
    credit_col = mapping.get("credit")
    debit = _parse_decimal(raw.get(debit_col)) if debit_col else None
    credit = _parse_decimal(raw.get(credit_col)) if credit_col else None

    if debit is not None and debit != 0 and credit is not None and credit != 0:
        # Ambiguous; treat as invalid.
        return None
    if debit is not None and debit != 0:
        return -abs(debit)
    if credit is not None and credit != 0:
        return abs(credit)

    return None


class ExcelParser:
    """Stateful .xlsx parser for import jobs."""

    def __init__(
        self,
        content: bytes,
        mapping_hint: Optional[ColumnMapping] = None,
        sheet_name: Optional[str] = None,
    ):
        self.content = content
        self.mapping_hint = mapping_hint or ColumnMapping()
        self.sheet_name = sheet_name

    def parse(self) -> dict[str, Any]:
        """Parse the workbook and return a structured result.

        Returns:
            {
                "headers": [...],
                "sheet_name": str,
                "mapping": {field: header, ...},
                "rows": [ParsedExcelRow, ...],
                "total_rows": int,
                "valid_rows": int,
                "invalid_rows": int,
            }
        """
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(self.content),
                read_only=True,
                data_only=True,
            )
        except (InvalidFileException, BadZipFile, KeyError, OSError) as exc:
            raise ExcelParseError(
                "Could not read the uploaded file as an Excel (.xlsx) workbook"
            ) from exc

        try:
            if self.sheet_name:
                if self.sheet_name not in workbook.sheetnames:
                    raise ExcelParseError(
                        f"Worksheet '{self.sheet_name}' not found. "
                        f"Available worksheets: {', '.join(workbook.sheetnames)}"
                    )
                worksheet = workbook[self.sheet_name]
            else:
                worksheet = workbook.worksheets[0]

            headers, data_rows = self._extract_header_and_rows(worksheet)
            sheet_title = worksheet.title
        finally:
            workbook.close()

        if headers is None:
            raise ExcelParseError("No header row found in the selected worksheet")

        mapping = _detect_mapping(headers, self.mapping_hint)

        rows: list[ParsedExcelRow] = []
        row_number = 1  # header is row 1

        for raw_values in data_rows:
            row_number += 1
            raw_typed = {
                headers[i]: (raw_values[i] if i < len(raw_values) else None)
                for i in range(len(headers))
            }
            # Skip completely blank rows.
            if not any(v is not None and str(v).strip() for v in raw_typed.values()):
                continue

            raw_json = {k: _json_safe(v) for k, v in raw_typed.items()}
            parsed_row = ParsedExcelRow(row_number, raw_json)
            self._parse_row(parsed_row, raw_typed, mapping)
            rows.append(parsed_row)

        valid_rows = sum(1 for r in rows if r.status == "valid")
        invalid_rows = len(rows) - valid_rows

        return {
            "headers": headers,
            "sheet_name": sheet_title,
            "mapping": mapping,
            "rows": rows,
            "total_rows": len(rows),
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
        }

    def _extract_header_and_rows(self, worksheet) -> tuple[Optional[list[str]], list[tuple]]:
        """Find the first non-blank row as the header; skip leading blank rows."""
        row_iter = worksheet.iter_rows(values_only=True)
        header_row: Optional[tuple] = None
        for row in row_iter:
            if row is None:
                continue
            if any(c is not None and str(c).strip() != "" for c in row):
                header_row = row
                break

        if header_row is None:
            return None, []

        headers: list[str] = []
        for i, cell in enumerate(header_row):
            if cell is None or str(cell).strip() == "":
                headers.append(f"Column {i + 1}")
            else:
                headers.append(str(cell).strip())

        data_rows = list(row_iter)
        return headers, data_rows

    def _parse_row(
        self, parsed_row: ParsedExcelRow, raw_typed: dict[str, Any], mapping: dict[str, str]
    ) -> None:
        parsed: dict[str, Any] = {}

        # Date
        date_col = mapping.get("date")
        if date_col:
            parsed_date = _cell_to_date(raw_typed.get(date_col))
            if parsed_date:
                parsed["date"] = parsed_date.isoformat()
            else:
                parsed_row.add_error(f"Could not parse date from '{date_col}'")
        else:
            parsed_row.add_error("No date column detected")

        # Description
        desc_col = mapping.get("description")
        if desc_col:
            value = raw_typed.get(desc_col)
            description = str(value).strip() if value is not None else ""
            if description:
                parsed["description"] = description
            else:
                parsed_row.add_error("Description is empty")
        else:
            parsed_row.add_error("No description column detected")

        # Amount handling: single amount, debit/credit columns, or negative amount.
        amount = _resolve_amount(raw_typed, mapping)
        if amount is None:
            parsed_row.add_error("Could not resolve a valid non-zero amount")
        else:
            parsed["amount"] = str(amount)
            parsed["amount_decimal"] = str(amount)
            if amount < 0:
                parsed["transaction_type"] = "expense"
            elif amount > 0:
                parsed["transaction_type"] = "income"
            else:
                parsed_row.add_error("Amount must be non-zero")

        # Explicit transaction type override.
        type_col = mapping.get("transaction_type")
        if type_col:
            txn_type = str(raw_typed.get(type_col, "")).strip().lower()
            if txn_type in ("expense", "debit", "out"):
                parsed["transaction_type"] = "expense"
            elif txn_type in ("income", "credit", "in"):
                parsed["transaction_type"] = "income"

        # Account / category columns (kept as text for later lookup).
        for field in ("account", "category", "reference", "currency", "balance"):
            col = mapping.get(field)
            if col and raw_typed.get(col) is not None:
                value = raw_typed.get(col)
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                parsed[field] = str(value).strip()

        parsed_row.parsed_data = parsed
        parsed_row.duplicate_key = _build_duplicate_key(parsed)


def parse_excel_import(
    content: bytes,
    mapping_hint: Optional[ColumnMapping] = None,
    sheet_name: Optional[str] = None,
) -> dict[str, Any]:
    """Convenience function to parse .xlsx content."""
    parser = ExcelParser(content, mapping_hint, sheet_name=sheet_name)
    return parser.parse()


def compute_excel_hash(content: bytes) -> str:
    """Return a SHA-256 hash of the raw file bytes for duplicate-file detection."""
    return hashlib.sha256(content).hexdigest()
